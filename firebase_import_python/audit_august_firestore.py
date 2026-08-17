from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import firebase_admin
import openpyxl
from firebase_admin import credentials, firestore

BASE_DIR = Path(__file__).resolve().parent
SERVICE_ACCOUNT_FILE = BASE_DIR / "serviceAccountKey.json"
EXCEL_FILE = BASE_DIR / "data_stok_beras_2026_08_agustus.xlsx"
REPORT_FILE = BASE_DIR / "audit_august_2026_report.txt"
YEAR = 2026
MONTH = 8
LOCAL_TZ = ZoneInfo("Asia/Jakarta")

COL_PRODUCTS = "products"
COL_BATCHES = "batches"
COL_TRANSACTIONS = "transactions"
COL_LOCATIONS = "storage_locations"
COL_COUNTERS = "counters"


@dataclass(frozen=True)
class Key:
    date: str
    product_id: str
    type: str


def slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"^_+|_+$", "", text)


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def parse_number(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return round(value)
    text = str(value).strip().replace(" ", "").replace(".", "").replace(",", ".")
    return round(float(text))


def parse_excel_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        from datetime import timedelta
        return datetime(1899, 12, 30) + timedelta(days=int(value))
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return datetime.fromisoformat(text)


def local_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if hasattr(value, "to_datetime"):
        value = value.to_datetime()
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        # Historical Python imports in this project used naive datetimes.
        # Treat their clock fields as Jakarta-local for audit grouping.
        return value.replace(tzinfo=LOCAL_TZ)
    return value.astimezone(LOCAL_TZ)


def date_str(value: datetime) -> str:
    return f"{value.year:04d}-{value.month:02d}-{value.day:02d}"


def load_excel_expected():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"File Excel tidak ditemukan: {EXCEL_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    try:
        ws = wb["Data Stok"] if "Data Stok" in wb.sheetnames else wb[wb.sheetnames[0]]
        raw_headers = [cell.value for cell in ws[1]]
        headers = [normalize_header(v) for v in raw_headers]
        aliases = {
            "date": ["tanggal", "date", "tgl"],
            "product": ["merkjenisberas", "merkberas", "jenisberas", "produk", "namaproduk", "barang"],
            "stock_in": ["stokmasuk", "stockin", "masuk", "qtymasuk", "jumlahmasuk"],
            "stock_out": ["stokkeluar", "stockout", "keluar", "qtykeluar", "jumlahkeluar"],
        }
        idx = {}
        for key, names in aliases.items():
            for name in names:
                if name in headers:
                    idx[key] = headers.index(name)
                    break
        missing = [k for k in aliases if k not in idx]
        if missing:
            raise RuntimeError(f"Kolom Excel tidak ditemukan: {missing}; header={raw_headers}")

        expected = defaultdict(int)
        rows = 0
        product_names = {}
        for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v is not None for v in values):
                continue
            dt = parse_excel_date(values[idx["date"]])
            if dt.year != YEAR or dt.month != MONTH:
                raise RuntimeError(f"Baris {row_no}: tanggal bukan Agustus 2026: {dt}")
            product_name = str(values[idx["product"]] or "").strip()
            product_id = slugify(product_name)
            stock_in = parse_number(values[idx["stock_in"]])
            stock_out = parse_number(values[idx["stock_out"]])
            if stock_in > 0:
                expected[Key(date_str(dt), product_id, "stock_in")] += stock_in
                rows += 1
            if stock_out > 0:
                expected[Key(date_str(dt), product_id, "stock_out")] += stock_out
                rows += 1
            product_names[product_id] = product_name
        return dict(expected), rows, product_names
    finally:
        wb.close()


def init_firestore():
    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(f"serviceAccountKey.json tidak ditemukan: {SERVICE_ACCOUNT_FILE}")
    try:
        firebase_admin.get_app()
    except ValueError:
        firebase_admin.initialize_app(credentials.Certificate(str(SERVICE_ACCOUNT_FILE)))
    return firestore.client()


def load_collection(db, name: str):
    result = {}
    for snap in db.collection(name).stream():
        result[snap.id] = snap.to_dict() or {}
    return result


def transaction_date(data: dict[str, Any]) -> datetime | None:
    # createdAt is the canonical field used by this project; keep fallbacks for audit only.
    for field in ("createdAt", "transactionDate", "date"):
        dt = local_datetime(data.get(field))
        if dt is not None:
            return dt
    return None


def format_qty(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def main():
    lines: list[str] = []

    def out(text=""):
        print(text)
        lines.append(text)

    out("AUDIT FIRESTORE AGUSTUS 2026 — READ ONLY")
    out("========================================")
    out("Script ini TIDAK melakukan set/update/delete pada Firestore.")
    out()

    expected, expected_rows, excel_names = load_excel_expected()
    out(f"Excel Agustus: {expected_rows} baris transaksi valid")
    out(f"Excel total stok masuk : {format_qty(sum(v for k, v in expected.items() if k.type == 'stock_in'))}")
    out(f"Excel total stok keluar: {format_qty(sum(v for k, v in expected.items() if k.type == 'stock_out'))}")
    out()

    out("Membaca Firestore...")
    db = init_firestore()
    products = load_collection(db, COL_PRODUCTS)
    batches = load_collection(db, COL_BATCHES)
    transactions = load_collection(db, COL_TRANSACTIONS)
    locations = load_collection(db, COL_LOCATIONS)
    counters = load_collection(db, COL_COUNTERS)
    out(f"products={len(products)}, batches={len(batches)}, transactions={len(transactions)}, locations={len(locations)}, counters={len(counters)}")
    out()

    august_docs = []
    actual = defaultdict(int)
    performed_by = Counter()
    id_prefixes = Counter()
    invalid_august_docs = []
    fingerprints = defaultdict(list)

    for doc_id, data in transactions.items():
        dt = transaction_date(data)
        if dt is None or dt.year != YEAR or dt.month != MONTH:
            continue
        tx_type = str(data.get("type") or "").strip().lower()
        if tx_type not in {"stock_in", "stock_out"}:
            continue
        product_id = str(data.get("productId") or "").strip()
        product_name = str(data.get("productName") or "").strip()
        qty = int(data.get("qty") or 0)
        batch_id = str(data.get("batchId") or "").strip()
        batch_code = str(data.get("batchCode") or "").strip()
        performer = str(data.get("performedBy") or data.get("createdBy") or "(kosong)").strip() or "(kosong)"
        performed_by[performer] += 1
        prefix = doc_id.split("-")[0] if "-" in doc_id else "random-id"
        id_prefixes[prefix] += 1

        if not product_id or qty <= 0:
            invalid_august_docs.append(doc_id)
            continue

        key = Key(date_str(dt), product_id, tx_type)
        actual[key] += qty
        august_docs.append((dt, doc_id, tx_type, product_id, product_name, qty, batch_id, batch_code, performer))
        fp = (date_str(dt), tx_type, product_id, qty, batch_id or batch_code)
        fingerprints[fp].append(doc_id)

    august_docs.sort(key=lambda x: (x[0], x[2], x[3], x[1]))
    out("RINGKASAN TRANSAKSI FIRESTORE AGUSTUS")
    out("-------------------------------------")
    out(f"Dokumen transaksi Agustus: {len(august_docs)}")
    out(f"  stock_in docs : {sum(1 for x in august_docs if x[2] == 'stock_in')}")
    out(f"  stock_out docs: {sum(1 for x in august_docs if x[2] == 'stock_out')}")
    out(f"Total qty stock_in : {format_qty(sum(x[5] for x in august_docs if x[2] == 'stock_in'))}")
    out(f"Total qty stock_out: {format_qty(sum(x[5] for x in august_docs if x[2] == 'stock_out'))}")
    if august_docs:
        out(f"Rentang tanggal: {august_docs[0][0].strftime('%d/%m/%Y')} s.d. {august_docs[-1][0].strftime('%d/%m/%Y')}")
    out()

    out("Sumber/performedBy transaksi Agustus:")
    for name, count in performed_by.most_common():
        out(f"- {name}: {count} dokumen")
    out()

    exact = []
    missing = []
    partial = []
    extra = []
    all_keys = sorted(set(expected) | set(actual), key=lambda k: (k.date, k.product_id, k.type))
    for key in all_keys:
        exp = expected.get(key, 0)
        act = actual.get(key, 0)
        if exp == act:
            exact.append((key, exp, act))
        elif exp > 0 and act == 0:
            missing.append((key, exp, act))
        elif exp > 0:
            partial.append((key, exp, act))
        else:
            extra.append((key, exp, act))

    out("PERBANDINGAN FIRESTORE VS EXCEL (tanggal + produk + tipe)")
    out("-------------------------------------------------------")
    out(f"Cocok persis : {len(exact)} kelompok")
    out(f"Belum ada    : {len(missing)} kelompok")
    out(f"Berbeda qty  : {len(partial)} kelompok")
    out(f"Ekstra DB    : {len(extra)} kelompok")
    out()

    def show_group(title, items, limit=100):
        out(title)
        if not items:
            out("- tidak ada")
            out()
            return
        for key, exp, act in items[:limit]:
            pname = excel_names.get(key.product_id) or str(products.get(key.product_id, {}).get("name") or key.product_id)
            out(f"- {key.date} | {pname} | {key.type} | Excel={exp} | Firestore={act} | selisih={act-exp:+d}")
        if len(items) > limit:
            out(f"... {len(items)-limit} kelompok lain tidak ditampilkan")
        out()

    show_group("BELUM ADA DI FIRESTORE:", missing)
    show_group("QTY BERBEDA:", partial)
    show_group("ADA DI FIRESTORE TAPI TIDAK ADA DI EXCEL:", extra)

    duplicates = [(fp, ids) for fp, ids in fingerprints.items() if len(ids) > 1]
    out("KANDIDAT DUPLIKAT SEMANTIK")
    out("-------------------------")
    if not duplicates:
        out("- tidak ditemukan fingerprint transaksi identik")
    else:
        for fp, ids in duplicates[:50]:
            out(f"- {fp}: {', '.join(ids)}")
    out()

    out("INTEGRITAS SALDO PRODUK vs BATCH AKTIF")
    out("-------------------------------------")
    active_batch_stock = defaultdict(int)
    active_batch_count = defaultdict(int)
    occupied_by_batch = defaultdict(list)
    august_received_batches = []

    for batch_id, data in batches.items():
        product_id = str(data.get("productId") or "").strip()
        remaining = int(data.get("remainingQty") or 0)
        status = str(data.get("status") or "").strip().lower()
        loc = str(data.get("storageLocation") or "").strip().upper()
        received = local_datetime(data.get("receivedAt"))
        if status == "active" and remaining > 0:
            active_batch_stock[product_id] += remaining
            active_batch_count[product_id] += 1
            if loc:
                occupied_by_batch[loc].append(batch_id)
        if received is not None and received.year == YEAR and received.month == MONTH:
            august_received_batches.append((received, batch_id, product_id, int(data.get("initialQty") or 0), remaining, status, loc))

    stock_mismatches = []
    for product_id, product in products.items():
        cached = int(product.get("totalStock") or 0)
        by_batch = active_batch_stock.get(product_id, 0)
        if cached != by_batch:
            stock_mismatches.append((product_id, cached, by_batch))

    if not stock_mismatches:
        out("✓ Semua products.totalStock sama dengan jumlah remainingQty batch aktif.")
    else:
        out(f"✗ Ditemukan {len(stock_mismatches)} produk tidak sinkron:")
        for product_id, cached, by_batch in stock_mismatches:
            pname = str(products.get(product_id, {}).get("name") or product_id)
            out(f"- {pname}: products.totalStock={cached}, batch aktif={by_batch}, selisih={cached-by_batch:+d}")
    out()

    duplicate_locations = {loc: ids for loc, ids in occupied_by_batch.items() if len(ids) > 1}
    out("INTEGRITAS LOKASI BATCH AKTIF")
    out("----------------------------")
    if not duplicate_locations:
        out("✓ Tidak ada satu lokasi yang dipakai oleh lebih dari satu batch aktif.")
    else:
        out("✗ Lokasi ganda ditemukan:")
        for loc, ids in duplicate_locations.items():
            out(f"- {loc}: {', '.join(ids)}")
    out()

    location_mismatches = []
    for loc, data in locations.items():
        is_occupied = data.get("isOccupied") is True
        batch_id = str(data.get("batchId") or "").strip()
        active_ids = occupied_by_batch.get(loc, [])
        if is_occupied:
            if len(active_ids) != 1 or (batch_id and active_ids and batch_id != active_ids[0]):
                location_mismatches.append((loc, is_occupied, batch_id, active_ids))
        else:
            if active_ids:
                location_mismatches.append((loc, is_occupied, batch_id, active_ids))

    if location_mismatches:
        out(f"✗ storage_locations tidak sinkron pada {len(location_mismatches)} lokasi:")
        for loc, is_occ, batch_id, active_ids in location_mismatches[:50]:
            out(f"- {loc}: isOccupied={is_occ}, batchId={batch_id or '-'}, batch aktif aktual={','.join(active_ids) or '-'}")
    else:
        out("✓ storage_locations konsisten dengan batch aktif.")
    out()

    august_received_batches.sort(key=lambda x: (x[0], x[1]))
    out("BATCH YANG DITERIMA PADA AGUSTUS")
    out("-------------------------------")
    out(f"Jumlah batch receivedAt Agustus: {len(august_received_batches)}")
    out(f"Total initialQty batch Agustus : {format_qty(sum(x[3] for x in august_received_batches))}")
    if august_received_batches:
        for rec, bid, pid, initial, remaining, status, loc in august_received_batches[:30]:
            pname = str(products.get(pid, {}).get("name") or pid)
            out(f"- {rec.strftime('%d/%m/%Y')} | {pname} | {bid} | initial={initial} | remaining={remaining} | {status} | {loc or '-'}")
        if len(august_received_batches) > 30:
            out(f"... {len(august_received_batches)-30} batch lain tidak ditampilkan")
    out()

    out("SAMPEL DOKUMEN TRANSAKSI AGUSTUS")
    out("--------------------------------")
    for dt, doc_id, tx_type, pid, pname, qty, batch_id, batch_code, performer in august_docs[:40]:
        out(f"- {dt.strftime('%d/%m/%Y %H:%M')} | {tx_type} | {pname or pid} | qty={qty} | batch={batch_code or batch_id or '-'} | by={performer} | id={doc_id}")
    if len(august_docs) > 40:
        out(f"... {len(august_docs)-40} transaksi lain tidak ditampilkan")
    out()

    out("KESIMPULAN OTOMATIS")
    out("-------------------")
    if len(actual) == 0:
        out("- Tidak ada transaksi Agustus di Firestore.")
    else:
        out(f"- Firestore sudah berisi {len(august_docs)} dokumen transaksi Agustus.")
    out(f"- Dari {len(expected)} kelompok Excel, {len(exact)} sudah cocok, {len(missing)} belum ada, {len(partial)} berbeda qty.")
    if extra:
        out(f"- Ada {len(extra)} kelompok transaksi Firestore yang tidak terdapat pada Excel Agustus.")
    if stock_mismatches or duplicate_locations or location_mismatches:
        out("- Ada masalah integritas state; JANGAN lakukan import tambahan sebelum diperbaiki/dipahami.")
    else:
        out("- State stok/batch/lokasi konsisten pada pemeriksaan dasar.")
    out("- Audit ini tidak mengubah data Firestore.")

    REPORT_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")
    out()
    print(f"Report tersimpan lokal: {REPORT_FILE.name}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\nAUDIT GAGAL")
        print("===========")
        print(exc)
        sys.exit(1)
