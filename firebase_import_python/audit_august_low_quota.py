from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import firebase_admin
from firebase_admin import credentials, firestore
from google.cloud.firestore_v1.base_query import FieldFilter
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
SERVICE_ACCOUNT_FILE = BASE_DIR / "serviceAccountKey.json"
EXCEL_FILE = BASE_DIR / "data_stok_beras_2026_08_agustus.xlsx"
REPORT_FILE = BASE_DIR / "audit_august_2026_low_quota_report.txt"

COL_PRODUCTS = "products"
COL_BATCHES = "batches"
COL_TRANSACTIONS = "transactions"
COL_LOCATIONS = "storage_locations"

YEAR = 2026
MONTH = 8
START = datetime(2026, 8, 1)
END = datetime(2026, 9, 1)

LOCATIONS = [
    *[f"A{i}" for i in range(1, 11)],
    *[f"B{i}" for i in range(1, 11)],
    *[f"C{i}" for i in range(1, 11)],
    *[f"D{i}" for i in range(1, 6)],
    *[f"X{i}" for i in range(1, 6)],
]


def initialize_firestore():
    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(f"serviceAccountKey.json tidak ditemukan: {SERVICE_ACCOUNT_FILE}")
    try:
        firebase_admin.get_app()
    except ValueError:
        firebase_admin.initialize_app(credentials.Certificate(str(SERVICE_ACCOUNT_FILE)))
    return firestore.client()


def norm_text(value: Any) -> str:
    return str(value or "").strip()


def parse_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    return int(float(value))


def product_id_from_name(name: str) -> str:
    return (
        name.strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
    )


def read_excel():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Data Stok"] if "Data Stok" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        headers[norm_text(cell.value).lower()] = idx

    required = {
        "tanggal",
        "merk/jenis beras",
        "stok masuk",
        "stok keluar",
    }
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"Kolom Excel tidak lengkap: {sorted(missing)}")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        raw_date = ws.cell(row_idx, headers["tanggal"]).value
        if not raw_date:
            continue

        if isinstance(raw_date, datetime):
            dt = raw_date
        else:
            raise RuntimeError(f"Tanggal baris {row_idx} tidak dikenali: {raw_date!r}")

        if dt.year != YEAR or dt.month != MONTH:
            raise RuntimeError(f"Excel berisi tanggal di luar Agustus 2026 pada baris {row_idx}: {dt}")

        product_name = norm_text(ws.cell(row_idx, headers["merk/jenis beras"]).value)
        stock_in = parse_int(ws.cell(row_idx, headers["stok masuk"]).value)
        stock_out = parse_int(ws.cell(row_idx, headers["stok keluar"]).value)

        if not product_name:
            continue
        if stock_in <= 0 and stock_out <= 0:
            continue

        product_id = product_id_from_name(product_name)
        date_key = dt.strftime("%Y-%m-%d")

        if stock_in > 0:
            rows.append((date_key, product_id, product_name, "stock_in", stock_in))
        if stock_out > 0:
            rows.append((date_key, product_id, product_name, "stock_out", stock_out))

    return rows


def get_documents(db, collection: str, ids):
    refs = [db.collection(collection).document(doc_id) for doc_id in sorted(set(ids))]
    result = {}
    for snap in db.get_all(refs):
        if snap.exists:
            result[snap.id] = snap.to_dict() or {}
    return result


def query_august_transactions(db):
    query = (
        db.collection(COL_TRANSACTIONS)
        .where(filter=FieldFilter("createdAt", ">=", START))
        .where(filter=FieldFilter("createdAt", "<", END))
    )
    result = []
    for snap in query.stream():
        data = snap.to_dict() or {}
        tx_type = norm_text(data.get("type")).lower()
        if tx_type in {"stock_in", "stock_out"}:
            result.append((snap.id, data))
    return result


def query_active_batches(db):
    query = db.collection(COL_BATCHES).where(
        filter=FieldFilter("status", "==", "active")
    )
    result = {}
    for snap in query.stream():
        data = snap.to_dict() or {}
        if parse_int(data.get("remainingQty")) > 0:
            result[snap.id] = data
    return result


def as_datetime(value: Any):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "to_datetime"):
        return value.to_datetime().replace(tzinfo=None)
    return None


def main():
    lines = []

    def out(text=""):
        print(text)
        lines.append(text)

    out("AUDIT AGUSTUS 2026 — LOW QUOTA / READ ONLY")
    out("==========================================")
    out("Script ini tidak melakukan set/update/delete.")
    out()

    excel_rows = read_excel()
    excel_agg = defaultdict(int)
    product_names = {}
    product_ids = set()

    for date_key, product_id, product_name, tx_type, qty in excel_rows:
        excel_agg[(date_key, product_id, tx_type)] += qty
        product_names[product_id] = product_name
        product_ids.add(product_id)

    out(f"Excel groups       : {len(excel_agg)}")
    out(f"Excel total masuk  : {sum(v for (d,p,t), v in excel_agg.items() if t == 'stock_in')}")
    out(f"Excel total keluar : {sum(v for (d,p,t), v in excel_agg.items() if t == 'stock_out')}")
    out()

    db = initialize_firestore()

    products = get_documents(db, COL_PRODUCTS, product_ids)
    active_batches = query_active_batches(db)
    august_transactions = query_august_transactions(db)
    location_docs = get_documents(db, COL_LOCATIONS, LOCATIONS)

    out("READ FIRESTORE HEMAT QUOTA")
    out("--------------------------")
    out(f"products dataset       : {len(products)}")
    out(f"active batches         : {len(active_batches)}")
    out(f"transactions Agustus   : {len(august_transactions)}")
    out(f"storage locations read : {len(location_docs)}")
    out()

    missing_products = sorted(product_ids - set(products))
    if missing_products:
        out("PRODUK DATASET TIDAK DITEMUKAN")
        for pid in missing_products:
            out(f"- {pid}")
        raise RuntimeError("Ada produk dataset yang tidak ditemukan di Firestore.")

    # Aggregate Firestore transactions by date + product + type.
    db_agg = defaultdict(int)
    stock_in_docs = 0
    stock_out_docs = 0

    for tx_id, data in august_transactions:
        dt = as_datetime(data.get("createdAt"))
        if not dt:
            continue
        product_id = norm_text(data.get("productId"))
        tx_type = norm_text(data.get("type")).lower()
        qty = parse_int(data.get("qty"))
        if tx_type == "stock_in":
            stock_in_docs += 1
        elif tx_type == "stock_out":
            stock_out_docs += 1
        else:
            continue
        db_agg[(dt.strftime("%Y-%m-%d"), product_id, tx_type)] += qty

    keys = sorted(set(excel_agg) | set(db_agg))
    matches = []
    missing = []
    different = []
    extras = []

    for key in keys:
        excel_qty = excel_agg.get(key, 0)
        db_qty = db_agg.get(key, 0)
        if excel_qty == db_qty and excel_qty > 0:
            matches.append((key, excel_qty, db_qty))
        elif excel_qty > 0 and db_qty == 0:
            missing.append((key, excel_qty, db_qty))
        elif excel_qty == 0 and db_qty > 0:
            extras.append((key, excel_qty, db_qty))
        else:
            different.append((key, excel_qty, db_qty))

    out("PERBANDINGAN EXCEL vs FIRESTORE")
    out("------------------------------")
    out(f"Cocok persis : {len(matches)}")
    out(f"Belum ada    : {len(missing)}")
    out(f"Berbeda qty  : {len(different)}")
    out(f"Ekstra DB    : {len(extras)}")
    out(f"stock_in docs Firestore  : {stock_in_docs}")
    out(f"stock_out docs Firestore : {stock_out_docs}")
    out(f"Total stock_in Firestore : {sum(v for (d,p,t), v in db_agg.items() if t == 'stock_in')}")
    out(f"Total stock_out Firestore: {sum(v for (d,p,t), v in db_agg.items() if t == 'stock_out')}")
    out()

    if different:
        out("QTY BERBEDA:")
        for (date_key, pid, tx_type), excel_qty, db_qty in different[:30]:
            out(f"- {date_key} | {product_names.get(pid, pid)} | {tx_type} | Excel={excel_qty} | DB={db_qty}")
        if len(different) > 30:
            out(f"... {len(different)-30} lainnya")
        out()

    if extras:
        out("EKSTRA FIRESTORE:")
        for (date_key, pid, tx_type), excel_qty, db_qty in extras[:30]:
            out(f"- {date_key} | {product_names.get(pid, pid)} | {tx_type} | DB={db_qty}")
        if len(extras) > 30:
            out(f"... {len(extras)-30} lainnya")
        out()

    # Product stock vs active batch remaining qty.
    batch_stock = defaultdict(int)
    occupied = {}
    duplicate_locations = []

    for batch_id, data in active_batches.items():
        pid = norm_text(data.get("productId"))
        remaining = parse_int(data.get("remainingQty"))
        loc = norm_text(data.get("storageLocation")).upper()

        if pid in product_ids:
            batch_stock[pid] += remaining

        if loc:
            if loc in occupied:
                duplicate_locations.append((loc, occupied[loc], batch_id))
            else:
                occupied[loc] = batch_id

    stock_mismatch = []
    for pid in sorted(product_ids):
        cached = parse_int(products[pid].get("totalStock"))
        actual = batch_stock.get(pid, 0)
        if cached != actual:
            stock_mismatch.append((pid, cached, actual))

    out("INTEGRITAS products.totalStock vs BATCH AKTIF")
    out("---------------------------------------------")
    if stock_mismatch:
        for pid, cached, actual in stock_mismatch:
            out(f"✗ {product_names.get(pid, pid)}: product={cached}, batch aktif={actual}")
    else:
        out("✓ Semua produk dataset konsisten.")
    out()

    out("INTEGRITAS LOKASI")
    out("----------------")
    if duplicate_locations:
        for loc, a, b in duplicate_locations:
            out(f"✗ Lokasi {loc} dipakai dua batch aktif: {a}, {b}")
    else:
        out("✓ Tidak ada lokasi ganda pada batch aktif.")

    location_mismatch = []
    for loc in LOCATIONS:
        loc_doc = location_docs.get(loc, {})
        expected_batch = occupied.get(loc)
        db_occupied = bool(loc_doc.get("isOccupied"))
        db_batch = norm_text(loc_doc.get("batchId"))

        if expected_batch:
            if not db_occupied or db_batch != expected_batch:
                location_mismatch.append((loc, expected_batch, db_batch or "-"))
        else:
            if db_occupied:
                location_mismatch.append((loc, "-", db_batch or "?"))

    if location_mismatch:
        for loc, expected, actual in location_mismatch:
            out(f"✗ {loc}: expected={expected}, storage_locations={actual}")
    else:
        out("✓ storage_locations konsisten dengan batch aktif.")
    out()

    success = (
        len(matches) == len(excel_agg)
        and not missing
        and not different
        and not extras
        and not stock_mismatch
        and not duplicate_locations
        and not location_mismatch
    )

    out("KESIMPULAN")
    out("----------")
    if success:
        out("✓ AUDIT LULUS: data Agustus Firestore cocok dengan Excel dan state stok konsisten.")
    else:
        out("✗ AUDIT BELUM LULUS: lihat bagian perbandingan/integritas di atas.")

    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")
    out()
    out(f"Report lokal: {REPORT_FILE.name}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print()
        print("AUDIT GAGAL")
        print("===========")
        print(exc)
