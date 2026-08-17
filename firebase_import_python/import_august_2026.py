from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import firebase_admin
import openpyxl
from firebase_admin import credentials, firestore


BASE_DIR = Path(__file__).resolve().parent
SERVICE_ACCOUNT_FILE = BASE_DIR / "serviceAccountKey.json"
EXCEL_FILE = "data_stok_beras_2026_08_agustus.xlsx"
PREFERRED_SHEET = "Data Stok"
EXPECTED_YEAR = 2026
EXPECTED_MONTH = 8
IMPORT_LOG_ID = "stock_august_2026"
IMPORT_USER_ID = "august_import"
IMPORT_USER_NAME = "Import Data Agustus"

COL_PRODUCTS = "products"
COL_BATCHES = "batches"
COL_TRANSACTIONS = "transactions"
COL_COUNTERS = "counters"
COL_LOCATIONS = "storage_locations"
COL_IMPORT_LOGS = "import_logs"

LOCATIONS = [
    *[f"A{number}" for number in range(1, 11)],
    *[f"B{number}" for number in range(1, 11)],
    *[f"C{number}" for number in range(1, 11)],
    *[f"D{number}" for number in range(1, 6)],
    *[f"X{number}" for number in range(1, 6)],
]

MAX_ATOMIC_WRITES = 490


@dataclass
class RowData:
    row_number: int
    row_key: str
    date: datetime
    product_id: str
    product_name: str
    stock_in: int
    stock_out: int
    unit: str
    notes: str


@dataclass
class BatchState:
    batch_id: str
    product_id: str
    product_name: str
    product_code: str
    received_at: datetime
    created_at: datetime
    initial_qty: int
    remaining_qty: int
    unit: str
    storage_location: str
    original_data: dict[str, Any]
    is_new: bool = False

    @property
    def batch_code(self) -> str:
        return str(self.original_data.get("batchCode") or self.batch_id).strip()


# ---------------------------------------------------------------------------
# Basic helpers
# ---------------------------------------------------------------------------

def initialize_firestore():
    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(
            "serviceAccountKey.json tidak ditemukan.\n"
            f"Letakkan file tersebut di folder yang sama dengan script:\n{SERVICE_ACCOUNT_FILE}"
        )

    try:
        firebase_admin.get_app()
    except ValueError:
        credential = credentials.Certificate(str(SERVICE_ACCOUNT_FILE))
        firebase_admin.initialize_app(credential)

    return firestore.client()


def slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"^_+|_+$", "", text)


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def parse_number(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return round(value)

    text = str(value).strip().replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return round(float(text))
    except ValueError as error:
        raise ValueError(f"angka tidak valid: {value}") from error


def parse_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)

    if isinstance(value, (int, float)):
        # Excel serial date
        from datetime import timedelta

        result = datetime(1899, 12, 30) + timedelta(days=int(value))
        return datetime(result.year, result.month, result.day)

    text = str(value or "").strip()
    if not text:
        raise ValueError("tanggal kosong")

    for date_format in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %m %Y"):
        try:
            result = datetime.strptime(text, date_format)
            return datetime(result.year, result.month, result.day)
        except ValueError:
            pass

    try:
        result = datetime.fromisoformat(text)
        return datetime(result.year, result.month, result.day)
    except ValueError as error:
        raise ValueError(f"format tanggal tidak valid: {value}") from error


def as_datetime(value: Any) -> datetime:
    if value is None:
        return datetime(1970, 1, 1)
    if isinstance(value, datetime):
        # Firestore may return timezone-aware datetimes. Sorting aware vs naive is invalid,
        # so strip tzinfo while preserving the represented clock fields.
        return value.replace(tzinfo=None)
    if hasattr(value, "to_datetime"):
        converted = value.to_datetime()
        return converted.replace(tzinfo=None)
    raise ValueError(f"Timestamp Firestore tidak dikenali: {value!r}")


def with_time(value: datetime, hour: int, minute: int = 0) -> datetime:
    return datetime(value.year, value.month, value.day, hour, minute, 0)


def display_date(value: datetime) -> str:
    return f"{value.day:02d}/{value.month:02d}/{value.year}"


def date_key(value: datetime) -> str:
    return f"{value.year}{value.month:02d}{value.day:02d}"


def extract_sequence(batch_code: str) -> int:
    parts = str(batch_code or "").strip().split("-")
    if not parts:
        return 0
    try:
        return int(parts[-1])
    except ValueError:
        return 0


def counter_id(product_id: str) -> str:
    return f"batch_sequence_{slugify(product_id) or 'unknown'}"


def location_zone(location: str) -> str:
    return "backup" if location.upper().startswith("X") else "main"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def get_sheet(workbook):
    if PREFERRED_SHEET in workbook.sheetnames:
        return workbook[PREFERRED_SHEET]
    return workbook[workbook.sheetnames[0]]


def build_header_index(sheet) -> dict[str, int]:
    raw_headers = [cell.value for cell in sheet[1]]
    headers = [normalize_header(value) for value in raw_headers]

    aliases = {
        "date": ["tanggal", "date", "tgl"],
        "product": [
            "merkjenisberas",
            "merkberas",
            "jenisberas",
            "produk",
            "namaproduk",
            "barang",
        ],
        "stock_in": ["stokmasuk", "stockin", "masuk", "qtymasuk", "jumlahmasuk"],
        "stock_out": ["stokkeluar", "stockout", "keluar", "qtykeluar", "jumlahkeluar"],
        "unit": ["satuan", "unit"],
        "notes": ["catatan", "notes", "keterangan"],
    }

    result: dict[str, int] = {}
    for key, names in aliases.items():
        for name in names:
            if name in headers:
                result[key] = headers.index(name)
                break

    required = ["date", "product", "stock_in", "stock_out"]
    missing = [key for key in required if key not in result]
    if missing:
        readable = [str(value or "").strip() for value in raw_headers]
        raise ValueError(
            f"Kolom wajib tidak ditemukan: {missing}. Header terbaca: {readable}"
        )

    return result


def read_august_rows() -> list[RowData]:
    path = BASE_DIR / EXCEL_FILE
    if not path.exists():
        raise FileNotFoundError(
            f"File Excel Agustus tidak ditemukan:\n{path}\n"
            "Pastikan file Excel berada di folder yang sama dengan script."
        )

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = get_sheet(workbook)
        index = build_header_index(sheet)
        source_slug = slugify(path.stem)
        rows: list[RowData] = []

        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            raw_date = values[index["date"]]
            raw_product = values[index["product"]]
            raw_stock_in = values[index["stock_in"]]
            raw_stock_out = values[index["stock_out"]]

            if (
                raw_date is None
                and raw_product is None
                and raw_stock_in is None
                and raw_stock_out is None
            ):
                continue

            try:
                item_date = parse_date(raw_date)
                stock_in = parse_number(raw_stock_in)
                stock_out = parse_number(raw_stock_out)
            except Exception as error:
                raise ValueError(
                    f"{EXCEL_FILE} baris {row_number}: {error}"
                ) from error

            if item_date.year != EXPECTED_YEAR or item_date.month != EXPECTED_MONTH:
                raise ValueError(
                    f"{EXCEL_FILE} baris {row_number}: tanggal "
                    f"{display_date(item_date)} bukan Agustus 2026"
                )

            if stock_in < 0 or stock_out < 0:
                raise ValueError(
                    f"{EXCEL_FILE} baris {row_number}: stok tidak boleh negatif"
                )

            if stock_in > 0 and stock_out > 0:
                raise ValueError(
                    f"{EXCEL_FILE} baris {row_number}: satu baris tidak boleh "
                    "berisi stok masuk dan stok keluar sekaligus"
                )

            if stock_in == 0 and stock_out == 0:
                continue

            product_id = slugify(raw_product)
            if not product_id:
                raise ValueError(
                    f"{EXCEL_FILE} baris {row_number}: nama produk kosong"
                )

            unit = "Karung"
            if "unit" in index:
                unit = str(values[index["unit"]] or "Karung").strip() or "Karung"

            notes = ""
            if "notes" in index:
                notes = str(values[index["notes"]] or "").strip()

            rows.append(
                RowData(
                    row_number=row_number,
                    row_key=f"{source_slug}_{row_number}",
                    date=item_date,
                    product_id=product_id,
                    product_name=str(raw_product or "").strip(),
                    stock_in=stock_in,
                    stock_out=stock_out,
                    unit=unit,
                    notes=notes,
                )
            )

        rows.sort(key=lambda item: (item.date, item.row_number))
        if not rows:
            raise ValueError("Tidak ada data Agustus yang valid pada file Excel")

        return rows
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Search keywords, following the app/importer's existing convention
# ---------------------------------------------------------------------------

def add_keyword(result: set[str], value: Any) -> None:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text or len(text) > 250:
        return

    result.add(text)
    compact = text.replace(" ", "")
    if len(compact) >= 2:
        result.add(compact)

    for word in text.split()[:30]:
        result.add(word)
        maximum_length = min(len(word), 24)
        for length in range(2, maximum_length + 1):
            result.add(word[:length])


def build_search_keywords(batch: dict[str, Any]) -> list[str]:
    result: set[str] = set()
    for value in (
        batch.get("id"),
        batch.get("batchCode"),
        batch.get("qrCodeValue"),
        batch.get("productId"),
        batch.get("productName"),
        batch.get("productCode"),
        batch.get("storageLocation"),
        batch.get("createdBy"),
        batch.get("createdByName"),
        batch.get("unit"),
        batch.get("initialQty"),
        batch.get("remainingQty"),
    ):
        add_keyword(result, value)

    remaining = int(batch.get("remainingQty") or 0)
    status = str(batch.get("status") or "").strip().lower()
    if status == "active" and remaining > 0:
        for value in ("active", "aktif", "batch aktif"):
            add_keyword(result, value)
    else:
        for value in ("empty", "depleted", "inactive", "habis", "tidak aktif", "batch habis"):
            add_keyword(result, value)

    received_at = as_datetime(batch.get("receivedAt"))
    month_names = [
        "januari",
        "februari",
        "maret",
        "april",
        "mei",
        "juni",
        "juli",
        "agustus",
        "september",
        "oktober",
        "november",
        "desember",
    ]
    month_name = month_names[received_at.month - 1]
    for value in (
        display_date(received_at),
        f"{received_at.day:02d}-{received_at.month:02d}-{received_at.year}",
        f"{received_at.year}-{received_at.month:02d}-{received_at.day:02d}",
        f"{received_at.day} {month_name} {received_at.year}",
        month_name,
        received_at.year,
    ):
        add_keyword(result, value)

    # Notes are intentionally included, same behavior as the current app.
    notes = str(batch.get("notes") or "").strip()
    if notes:
        normalized_notes = re.sub(r"\s+", " ", notes.lower()).strip()
        if len(normalized_notes) <= 250:
            result.add(normalized_notes)
        for word in normalized_notes.split()[:30]:
            add_keyword(result, word)

    return sorted(keyword for keyword in result if keyword and len(keyword) <= 250)


# ---------------------------------------------------------------------------
# Read current Firestore state and validate it before simulating August
# ---------------------------------------------------------------------------

def load_collection(database, collection_name: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for snapshot in database.collection(collection_name).stream():
        result[snapshot.id] = snapshot.to_dict() or {}
    return result


def existing_august_transactions(transactions: dict[str, dict[str, Any]]) -> list[str]:
    found: list[str] = []
    for document_id, data in transactions.items():
        created_at = data.get("createdAt")
        if created_at is None:
            continue
        try:
            value = as_datetime(created_at)
        except Exception:
            continue

        transaction_type = str(data.get("type") or "").strip().lower()
        if (
            value.year == EXPECTED_YEAR
            and value.month == EXPECTED_MONTH
            and transaction_type in {"stock_in", "stock_out"}
        ):
            found.append(document_id)
    return found


def make_batch_state(document_id: str, data: dict[str, Any]) -> BatchState:
    return BatchState(
        batch_id=document_id,
        product_id=str(data.get("productId") or "").strip(),
        product_name=str(data.get("productName") or "").strip(),
        product_code=str(data.get("productCode") or "").strip().upper(),
        received_at=as_datetime(data.get("receivedAt")),
        created_at=as_datetime(data.get("createdAt") or data.get("receivedAt")),
        initial_qty=int(data.get("initialQty") or 0),
        remaining_qty=int(data.get("remainingQty") or 0),
        unit=str(data.get("unit") or "Karung").strip() or "Karung",
        storage_location=str(data.get("storageLocation") or "").strip().upper(),
        original_data=dict(data),
        is_new=False,
    )


def batch_sort_key(batch: BatchState):
    sequence = extract_sequence(batch.batch_code)
    return (
        batch.received_at,
        batch.created_at,
        sequence if sequence > 0 else 10**9,
        batch.batch_code.upper(),
        batch.batch_id.upper(),
    )


def validate_current_state(
    products: dict[str, dict[str, Any]],
    all_batches: dict[str, dict[str, Any]],
) -> tuple[dict[str, list[BatchState]], dict[str, str], dict[str, int]]:
    queues: dict[str, list[BatchState]] = {product_id: [] for product_id in products}
    occupied_locations: dict[str, str] = {}
    max_sequences: dict[str, int] = {product_id: 0 for product_id in products}

    for document_id, data in all_batches.items():
        product_id = str(data.get("productId") or "").strip()
        if not product_id:
            continue

        batch_code = str(data.get("batchCode") or document_id).strip()
        sequence = extract_sequence(batch_code)
        if product_id in max_sequences and sequence > max_sequences[product_id]:
            max_sequences[product_id] = sequence

        status = str(data.get("status") or "").strip().lower()
        remaining = int(data.get("remainingQty") or 0)
        if status != "active" or remaining <= 0:
            continue

        if product_id not in products:
            raise RuntimeError(
                f"Batch aktif {document_id} menunjuk productId yang tidak ada: {product_id}"
            )

        batch = make_batch_state(document_id, data)
        if not batch.storage_location:
            raise RuntimeError(f"Batch aktif {document_id} belum memiliki storageLocation")
        if batch.storage_location not in LOCATIONS:
            raise RuntimeError(
                f"Batch aktif {document_id} memiliki lokasi tidak valid: {batch.storage_location}"
            )
        if batch.storage_location in occupied_locations:
            raise RuntimeError(
                f"Lokasi {batch.storage_location} dipakai lebih dari satu batch aktif: "
                f"{occupied_locations[batch.storage_location]} dan {document_id}"
            )

        occupied_locations[batch.storage_location] = document_id
        queues[product_id].append(batch)

    for queue in queues.values():
        queue.sort(key=batch_sort_key)

    mismatches: list[str] = []
    for product_id, product in products.items():
        actual_batch_stock = sum(batch.remaining_qty for batch in queues.get(product_id, []))
        cached_product_stock = int(product.get("totalStock") or 0)
        if actual_batch_stock != cached_product_stock:
            mismatches.append(
                f"{product_id}: products.totalStock={cached_product_stock}, "
                f"jumlah batch aktif={actual_batch_stock}"
            )

    if mismatches:
        raise RuntimeError(
            "Saldo Firestore saat ini tidak konsisten. Import Agustus dibatalkan agar data lama "
            "tidak semakin rusak.\n- " + "\n- ".join(mismatches)
        )

    return queues, occupied_locations, max_sequences


# ---------------------------------------------------------------------------
# Simulate August entirely in memory. Nothing is written here.
# ---------------------------------------------------------------------------

def allocate_location(occupied_locations: dict[str, str]) -> str:
    for location in LOCATIONS:
        if location not in occupied_locations:
            return location
    raise RuntimeError("Tidak ada lokasi penyimpanan kosong untuk batch baru Agustus")


def state_to_batch_document(batch: BatchState) -> dict[str, Any]:
    data = dict(batch.original_data)
    data.update(
        {
            "id": batch.batch_id,
            "productId": batch.product_id,
            "productName": batch.product_name,
            "productCode": batch.product_code,
            "batchCode": batch.batch_code,
            "receivedAt": batch.received_at,
            "initialQty": batch.initial_qty,
            "remainingQty": batch.remaining_qty,
            "unit": batch.unit,
            "qrCodeValue": batch.batch_code,
            "status": "active" if batch.remaining_qty > 0 else "empty",
            "storageLocation": batch.storage_location,
            "updatedAt": firestore.SERVER_TIMESTAMP,
        }
    )
    if batch.is_new:
        data.setdefault("createdBy", IMPORT_USER_ID)
        data.setdefault("createdByName", IMPORT_USER_NAME)
        data.setdefault("createdAt", batch.created_at)
    data["searchKeywords"] = build_search_keywords(data)
    return data


def build_location_document(location: str, batch: BatchState | None) -> dict[str, Any]:
    if batch is None:
        return {
            "id": location,
            "location": location,
            "zone": location_zone(location),
            "isOccupied": False,
            "batchId": None,
            "batchCode": None,
            "productId": None,
            "productName": None,
            "remainingQty": 0,
            "occupiedAt": None,
            "releasedAt": firestore.SERVER_TIMESTAMP,
            "updatedAt": firestore.SERVER_TIMESTAMP,
        }

    return {
        "id": location,
        "location": location,
        "zone": location_zone(location),
        "isOccupied": True,
        "batchId": batch.batch_id,
        "batchCode": batch.batch_code,
        "productId": batch.product_id,
        "productName": batch.product_name,
        "remainingQty": batch.remaining_qty,
        "occupiedAt": batch.received_at,
        "releasedAt": None,
        "updatedAt": firestore.SERVER_TIMESTAMP,
    }


def simulate_august(
    rows: list[RowData],
    products: dict[str, dict[str, Any]],
    all_batches: dict[str, dict[str, Any]],
    counters: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    queues, occupied_locations, max_sequences = validate_current_state(products, all_batches)

    # Dataset is not allowed to silently create new master products.
    dataset_products = sorted({row.product_id for row in rows})
    missing_products = [product_id for product_id in dataset_products if product_id not in products]
    if missing_products:
        raise RuntimeError(
            "Produk pada Excel belum ada di collection products. Import dibatalkan:\n- "
            + "\n- ".join(missing_products)
        )

    # Continue sequence from the larger value between current counter and all existing batches.
    sequences: dict[str, int] = {}
    for product_id in products:
        existing_counter = counters.get(counter_id(product_id), {})
        counter_last = int(existing_counter.get("lastNumber") or 0)
        sequences[product_id] = max(counter_last, max_sequences.get(product_id, 0))

    product_stocks = {
        product_id: int(product.get("totalStock") or 0)
        for product_id, product in products.items()
    }

    changed_batches: dict[str, BatchState] = {}
    new_batches: set[str] = set()
    transactions: dict[str, dict[str, Any]] = {}
    touched_products: set[str] = set()
    touched_locations: set[str] = set()
    touched_counters: set[str] = set()

    stock_in_rows = 0
    stock_out_rows = 0
    stock_out_transaction_parts = 0

    for row in rows:
        product = products[row.product_id]
        is_active = product.get("isActive") is True
        is_deleted = product.get("isDeleted") is True
        if not is_active or is_deleted:
            raise RuntimeError(
                f"Produk {row.product_id} tidak aktif/dihapus sehingga tidak boleh dipakai"
            )

        product_code = str(product.get("code") or "").strip().upper()
        product_name = str(product.get("name") or row.product_name).strip()
        product_unit = str(product.get("unit") or row.unit or "Karung").strip() or "Karung"
        if not product_code:
            raise RuntimeError(f"Kode produk kosong pada Firestore: {row.product_id}")

        if row.stock_in > 0:
            stock_in_rows += 1
            touched_products.add(row.product_id)
            touched_counters.add(row.product_id)

            sequences[row.product_id] += 1
            sequence = sequences[row.product_id]
            batch_code = (
                f"BATCH-{date_key(row.date)}-{product_code}-{sequence:03d}"
            )

            if batch_code in all_batches or batch_code in new_batches:
                raise RuntimeError(f"Kode batch Agustus sudah ada: {batch_code}")

            location = allocate_location(occupied_locations)
            touched_locations.add(location)
            occupied_locations[location] = batch_code

            notes = (
                f"Import data Agustus 2026 dari {EXCEL_FILE}, tanggal "
                f"{display_date(row.date)}. {row.notes}"
            ).strip()

            batch_data = {
                "id": batch_code,
                "productId": row.product_id,
                "productName": product_name,
                "productCode": product_code,
                "batchCode": batch_code,
                "receivedAt": with_time(row.date, 8, 0),
                "initialQty": row.stock_in,
                "remainingQty": row.stock_in,
                "unit": product_unit,
                "qrCodeValue": batch_code,
                "status": "active",
                "storageLocation": location,
                "createdBy": IMPORT_USER_ID,
                "createdByName": IMPORT_USER_NAME,
                "notes": notes,
                "createdAt": with_time(row.date, 8, 0),
                "updatedAt": firestore.SERVER_TIMESTAMP,
            }

            batch = BatchState(
                batch_id=batch_code,
                product_id=row.product_id,
                product_name=product_name,
                product_code=product_code,
                received_at=with_time(row.date, 8, 0),
                created_at=with_time(row.date, 8, 0),
                initial_qty=row.stock_in,
                remaining_qty=row.stock_in,
                unit=product_unit,
                storage_location=location,
                original_data=batch_data,
                is_new=True,
            )
            queues[row.product_id].append(batch)
            queues[row.product_id].sort(key=batch_sort_key)
            changed_batches[batch.batch_id] = batch
            new_batches.add(batch.batch_id)

            product_stocks[row.product_id] += row.stock_in

            transaction_id = f"STOCK-IN-{batch_code}"
            transactions[transaction_id] = {
                "id": transaction_id,
                "type": "stock_in",
                "productId": row.product_id,
                "productName": product_name,
                "batchId": batch_code,
                "batchCode": batch_code,
                "qty": row.stock_in,
                "unit": product_unit,
                "performedBy": IMPORT_USER_ID,
                "performedByName": IMPORT_USER_NAME,
                "notes": notes,
                "createdAt": with_time(row.date, 8, 10),
            }

        if row.stock_out > 0:
            stock_out_rows += 1
            touched_products.add(row.product_id)
            remaining_out = row.stock_out
            part_number = 1

            while remaining_out > 0:
                queue = queues[row.product_id]
                current_batch = next(
                    (batch for batch in queue if batch.remaining_qty > 0), None
                )
                if current_batch is None:
                    raise RuntimeError(
                        "Stok keluar Agustus melebihi stok yang tersedia.\n"
                        f"Baris Excel: {row.row_number}\n"
                        f"Tanggal: {display_date(row.date)}\n"
                        f"Produk: {product_name}\n"
                        f"Kekurangan: {remaining_out} {product_unit}"
                    )

                allocated = min(remaining_out, current_batch.remaining_qty)
                before_qty = current_batch.remaining_qty
                current_batch.remaining_qty -= allocated
                remaining_out -= allocated
                product_stocks[row.product_id] -= allocated

                if product_stocks[row.product_id] < 0:
                    raise RuntimeError(f"Total stok {product_name} menjadi negatif")

                changed_batches[current_batch.batch_id] = current_batch
                touched_locations.add(current_batch.storage_location)

                after_qty = current_batch.remaining_qty
                if after_qty <= 0:
                    occupied_locations.pop(current_batch.storage_location, None)

                transaction_id = (
                    f"AUG-OUT-{row.row_key}-{part_number:03d}-{current_batch.batch_id}"
                )
                notes = (
                    f"Import stok keluar FIFO Agustus 2026 dari {EXCEL_FILE}, tanggal "
                    f"{display_date(row.date)}. {row.notes}"
                ).strip()

                transactions[transaction_id] = {
                    "id": transaction_id,
                    "type": "stock_out",
                    "productId": row.product_id,
                    "productName": product_name,
                    "batchId": current_batch.batch_id,
                    "batchCode": current_batch.batch_code,
                    "qty": allocated,
                    "unit": current_batch.unit,
                    "performedBy": IMPORT_USER_ID,
                    "performedByName": IMPORT_USER_NAME,
                    "notes": notes,
                    "storageLocation": current_batch.storage_location,
                    "remainingQtyBefore": before_qty,
                    "remainingQtyAfter": after_qty,
                    "totalStockAfter": product_stocks[row.product_id],
                    "createdAt": with_time(row.date, 15, min(part_number, 59)),
                }

                stock_out_transaction_parts += 1
                part_number += 1

    # Final occupied mapping must point to the active batch state after August simulation.
    active_by_location: dict[str, BatchState] = {}
    for queue in queues.values():
        for batch in queue:
            if batch.remaining_qty <= 0:
                continue
            location = batch.storage_location
            if location in active_by_location:
                raise RuntimeError(
                    f"Simulasi menghasilkan dua batch aktif pada lokasi {location}"
                )
            active_by_location[location] = batch

    # Prepare only documents that need to change.
    batch_documents = {
        batch_id: state_to_batch_document(batch)
        for batch_id, batch in changed_batches.items()
    }

    product_documents: dict[str, dict[str, Any]] = {}
    for product_id in touched_products:
        product_documents[product_id] = {
            "totalStock": product_stocks[product_id],
            "updatedAt": firestore.SERVER_TIMESTAMP,
        }

    counter_documents: dict[str, dict[str, Any]] = {}
    for product_id in touched_counters:
        product = products[product_id]
        document_id = counter_id(product_id)
        counter_documents[document_id] = {
            "id": document_id,
            "productId": product_id,
            "productCode": str(product.get("code") or "").strip().upper(),
            "lastNumber": sequences[product_id],
            "updatedAt": firestore.SERVER_TIMESTAMP,
        }

    location_documents: dict[str, dict[str, Any]] = {}
    for location in touched_locations:
        location_documents[location] = build_location_document(
            location, active_by_location.get(location)
        )

    total_in = sum(row.stock_in for row in rows)
    total_out = sum(row.stock_out for row in rows)
    opening_stock = sum(int(product.get("totalStock") or 0) for product in products.values())
    closing_stock = sum(product_stocks.values())
    expected_closing = opening_stock + total_in - total_out
    if closing_stock != expected_closing:
        raise RuntimeError(
            "Validasi saldo simulasi gagal.\n"
            f"Saldo awal: {opening_stock}\n"
            f"Masuk Agustus: {total_in}\n"
            f"Keluar Agustus: {total_out}\n"
            f"Seharusnya: {expected_closing}\n"
            f"Hasil simulasi: {closing_stock}"
        )

    return {
        "rows": len(rows),
        "stock_in_rows": stock_in_rows,
        "stock_out_rows": stock_out_rows,
        "stock_out_transaction_parts": stock_out_transaction_parts,
        "total_in": total_in,
        "total_out": total_out,
        "opening_stock": opening_stock,
        "closing_stock": closing_stock,
        "batch_documents": batch_documents,
        "new_batches_count": len(new_batches),
        "transactions": transactions,
        "product_documents": product_documents,
        "counter_documents": counter_documents,
        "location_documents": location_documents,
        "product_stocks": product_stocks,
    }


# ---------------------------------------------------------------------------
# Commit + verification
# ---------------------------------------------------------------------------

def print_summary(data: dict[str, Any]) -> None:
    print()
    print("RINGKASAN SIMULASI AGUSTUS 2026")
    print("===============================")
    print(f"Baris Excel diproses       : {data['rows']}")
    print(f"Baris stok masuk           : {data['stock_in_rows']}")
    print(f"Baris stok keluar          : {data['stock_out_rows']}")
    print(f"Batch baru                 : {data['new_batches_count']}")
    print(f"Dokumen batch diubah       : {len(data['batch_documents'])}")
    print(f"Dokumen transaksi dibuat   : {len(data['transactions'])}")
    print(f"  bagian stock_out FIFO    : {data['stock_out_transaction_parts']}")
    print(f"Produk diperbarui          : {len(data['product_documents'])}")
    print(f"Counter diperbarui         : {len(data['counter_documents'])}")
    print(f"Lokasi diperbarui          : {len(data['location_documents'])}")
    print(f"Total stok masuk           : {data['total_in']}")
    print(f"Total stok keluar          : {data['total_out']}")
    print(f"Saldo awal sebelum Agustus : {data['opening_stock']}")
    print(f"Saldo akhir sesudah Agustus: {data['closing_stock']}")

    print()
    print("SALDO PRODUK SETELAH AGUSTUS")
    print("============================")
    for product_id, stock in sorted(data["product_stocks"].items()):
        print(f"- {product_id:<20} {stock}")


def total_write_count(data: dict[str, Any]) -> int:
    return (
        len(data["batch_documents"])
        + len(data["transactions"])
        + len(data["product_documents"])
        + len(data["counter_documents"])
        + len(data["location_documents"])
        + 1  # import log
    )


def commit_august(database, data: dict[str, Any]) -> None:
    writes = total_write_count(data)
    if writes > MAX_ATOMIC_WRITES:
        raise RuntimeError(
            f"Import membutuhkan {writes} write. Batas aman script adalah "
            f"{MAX_ATOMIC_WRITES} agar seluruh import dapat dikirim atomik. "
            "Jangan lanjut; kirim output ini untuk penyesuaian script."
        )

    print()
    print(f"Menulis {writes} dokumen secara atomik ke Firestore...")
    batch = database.batch()

    for document_id, value in data["batch_documents"].items():
        ref = database.collection(COL_BATCHES).document(document_id)
        batch.set(ref, value, merge=True)

    for document_id, value in data["transactions"].items():
        ref = database.collection(COL_TRANSACTIONS).document(document_id)
        batch.set(ref, value)

    for document_id, value in data["product_documents"].items():
        ref = database.collection(COL_PRODUCTS).document(document_id)
        batch.set(ref, value, merge=True)

    for document_id, value in data["counter_documents"].items():
        ref = database.collection(COL_COUNTERS).document(document_id)
        batch.set(ref, value, merge=True)

    for document_id, value in data["location_documents"].items():
        ref = database.collection(COL_LOCATIONS).document(document_id)
        batch.set(ref, value, merge=True)

    import_log = {
        "id": IMPORT_LOG_ID,
        "description": "Incremental import data stok Agustus 2026",
        "periodStart": datetime(2026, 8, 1),
        "periodEnd": datetime(2026, 8, 31, 23, 59, 59),
        "sourceFiles": [EXCEL_FILE],
        "rowsProcessed": data["rows"],
        "stockInRows": data["stock_in_rows"],
        "stockOutRows": data["stock_out_rows"],
        "newBatchesCount": data["new_batches_count"],
        "transactionCount": len(data["transactions"]),
        "totalStockIn": data["total_in"],
        "totalStockOut": data["total_out"],
        "openingStock": data["opening_stock"],
        "closingStock": data["closing_stock"],
        "createdAt": firestore.SERVER_TIMESTAMP,
    }
    batch.set(
        database.collection(COL_IMPORT_LOGS).document(IMPORT_LOG_ID),
        import_log,
    )

    batch.commit()
    print("Commit Firestore selesai.")


def verify_after_commit(database, data: dict[str, Any]) -> None:
    errors: list[str] = []

    log = database.collection(COL_IMPORT_LOGS).document(IMPORT_LOG_ID).get()
    if not log.exists:
        errors.append("import log Agustus tidak ditemukan")

    for product_id, expected_stock in data["product_stocks"].items():
        product = database.collection(COL_PRODUCTS).document(product_id).get()
        if not product.exists:
            errors.append(f"produk hilang: {product_id}")
            continue
        saved_stock = int((product.to_dict() or {}).get("totalStock") or 0)
        if saved_stock != expected_stock:
            errors.append(
                f"stok {product_id}: aktual {saved_stock}, seharusnya {expected_stock}"
            )

    # Spot-check all deterministic transaction IDs created by this import.
    missing_transactions: list[str] = []
    for transaction_id in data["transactions"]:
        if not database.collection(COL_TRANSACTIONS).document(transaction_id).get().exists:
            missing_transactions.append(transaction_id)
            if len(missing_transactions) >= 5:
                break
    if missing_transactions:
        errors.append(
            "transaksi hasil import tidak ditemukan: " + ", ".join(missing_transactions)
        )

    if errors:
        raise RuntimeError("Verifikasi setelah commit gagal:\n- " + "\n- ".join(errors))

    print()
    print("IMPORT AGUSTUS BERHASIL DAN TERVERIFIKASI")
    print("========================================")
    print("Data Januari-Juli tidak dihapus atau di-import ulang.")
    print("Yang ditambahkan adalah transaksi/batch Agustus dan pembaruan saldo berjalan.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Incremental import data stok Agustus 2026 ke Firestore tanpa "
            "mengimpor ulang Januari-Juli"
        )
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Benar-benar menulis hasil simulasi Agustus ke Firestore",
    )
    arguments = parser.parse_args()

    print()
    print("MODE:", "COMMIT KE FIRESTORE" if arguments.commit else "DRY RUN")
    print("Data lama Januari-Juli TIDAK akan di-import ulang.")
    print()

    print(f"Membaca {EXCEL_FILE}...")
    rows = read_august_rows()
    print(f"  {len(rows)} baris valid")

    print("Menghubungkan ke Firestore untuk membaca saldo saat ini...")
    database = initialize_firestore()

    import_log = database.collection(COL_IMPORT_LOGS).document(IMPORT_LOG_ID).get()
    if import_log.exists:
        raise RuntimeError(
            "Import log stock_august_2026 sudah ada. Script dihentikan untuk "
            "mencegah data Agustus masuk dua kali."
        )

    products = load_collection(database, COL_PRODUCTS)
    all_batches = load_collection(database, COL_BATCHES)
    counters = load_collection(database, COL_COUNTERS)
    transactions_existing = load_collection(database, COL_TRANSACTIONS)

    august_existing = existing_august_transactions(transactions_existing)
    if august_existing:
        preview = "\n- ".join(august_existing[:10])
        raise RuntimeError(
            "Firestore sudah memiliki transaksi stock_in/stock_out bertanggal Agustus 2026. "
            "Import penuh Agustus dibatalkan agar tidak terjadi duplikasi.\n"
            f"Contoh dokumen:\n- {preview}\n"
            "Jika transaksi tersebut memang harus dipertahankan, jangan hapus sendiri; "
            "sesuaikan importer berdasarkan kondisi database."
        )

    data = simulate_august(
        rows=rows,
        products=products,
        all_batches=all_batches,
        counters=counters,
    )
    print_summary(data)

    writes = total_write_count(data)
    print()
    print(f"Total write yang direncanakan: {writes}")
    if writes > MAX_ATOMIC_WRITES:
        raise RuntimeError(
            f"Jumlah write {writes} melebihi batas aman atomik {MAX_ATOMIC_WRITES}."
        )

    if not arguments.commit:
        print()
        print("DRY RUN BERHASIL")
        print("================")
        print("Belum ada data yang ditulis ke Firestore.")
        print("Jika ringkasan di atas benar, jalankan:")
        print("python import_august_2026.py --commit")
        return

    commit_august(database, data)
    verify_after_commit(database, data)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProses dihentikan.")
        raise SystemExit(130)
    except Exception as error:
        print()
        print("IMPORT AGUSTUS GAGAL")
        print("====================")
        print(error)
        sys.exit(1)
