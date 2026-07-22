from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

import firebase_admin
import openpyxl
from firebase_admin import credentials, firestore


BASE_DIR = Path(__file__).resolve().parent

SERVICE_ACCOUNT_FILE = (
    BASE_DIR
    / "serviceAccountKey.json"
)

EXCEL_FILES = [
    (
        "data_stok_beras_2026_01_januari.xlsx",
        2026,
        1,
    ),
    (
        "data_stok_beras_2026_02_februari.xlsx",
        2026,
        2,
    ),
    (
        "data_stok_beras_2026_03_maret.xlsx",
        2026,
        3,
    ),
    (
        "data_stok_beras_2026_04_april.xlsx",
        2026,
        4,
    ),
    (
        "data_stok_beras_2026_05_mei.xlsx",
        2026,
        5,
    ),
    (
        "data_stok_beras_2026_06_juni.xlsx",
        2026,
        6,
    ),
    (
        "data_stok_beras_2026_07_juli.xlsx",
        2026,
        7,
    ),
]

PREFERRED_SHEET = "Data Stok"

IMPORT_LOG_ID = (
    "initial_stock_jan_jul_2026"
)

IMPORT_USER_ID = "initial_import"
IMPORT_USER_NAME = "Import Data Awal"

COL_PRODUCTS = "products"
COL_BATCHES = "batches"
COL_TRANSACTIONS = "transactions"
COL_COUNTERS = "counters"
COL_LOCATIONS = "storage_locations"
COL_IMPORT_LOGS = "import_logs"

DEFAULT_MINIMUM_STOCK = 25

LOCATIONS = [
    *[
        f"A{number}"
        for number in range(1, 11)
    ],
    *[
        f"B{number}"
        for number in range(1, 11)
    ],
    *[
        f"C{number}"
        for number in range(1, 11)
    ],
    *[
        f"D{number}"
        for number in range(1, 6)
    ],
    *[
        f"X{number}"
        for number in range(1, 6)
    ],
]

PRODUCT_CODES = {
    "beras_walet": "BR-001",
    "bmw": "BR-002",
    "cap_kembang": "BR-003",
    "ir_cap_mercy": "BR-004",
    "karya_makmur": "BR-005",
    "keraton": "BR-006",
    "pandan_wangi": "BR-007",
    "ramos": "BR-008",
    "rojo_lele": "BR-009",
    "sumber_raya": "BR-010",
}


@dataclass
class RowData:
    source_file: str
    source_order: int
    row_number: int
    row_key: str
    date: datetime
    product_id: str
    product_name: str
    product_code: str
    stock_in: int
    stock_out: int
    unit: str
    notes: str


@dataclass
class QueueBatch:
    batch_id: str
    remaining_qty: int
    location: str


def initialize_firestore():
    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(
            "serviceAccountKey.json "
            "tidak ditemukan:\n"
            f"{SERVICE_ACCOUNT_FILE}"
        )

    try:
        firebase_admin.get_app()
    except ValueError:
        credential = credentials.Certificate(
            str(SERVICE_ACCOUNT_FILE)
        )

        firebase_admin.initialize_app(
            credential
        )

    return firestore.client()


def slugify(value) -> str:
    text = str(
        value or ""
    ).strip().lower()

    text = re.sub(
        r"[^a-z0-9]+",
        "_",
        text,
    )

    return re.sub(
        r"^_+|_+$",
        "",
        text,
    )


def title_case(value) -> str:
    words = str(
        value or ""
    ).strip().split()

    result: list[str] = []

    for word in words:
        if (
            word.isupper()
            and len(word) <= 5
        ):
            result.append(
                word
            )
        else:
            result.append(
                word[:1].upper()
                + word[1:].lower()
            )

    return " ".join(
        result
    )


def normalize_header(value) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(
            value or ""
        ).strip().lower(),
    )


def parse_number(value) -> int:
    if value is None or value == "":
        return 0

    if isinstance(value, bool):
        return 0

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return round(value)

    text = (
        str(value)
        .strip()
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )

    try:
        return round(
            float(text)
        )
    except ValueError as error:
        raise ValueError(
            f"angka tidak valid: {value}"
        ) from error


def parse_date(value) -> datetime:
    if isinstance(value, datetime):
        return datetime(
            value.year,
            value.month,
            value.day,
        )

    if isinstance(value, (int, float)):
        result = (
            datetime(1899, 12, 30)
            + timedelta(
                days=int(value)
            )
        )

        return datetime(
            result.year,
            result.month,
            result.day,
        )

    text = str(
        value or ""
    ).strip()

    if not text:
        raise ValueError(
            "tanggal kosong"
        )

    formats = (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d %m %Y",
    )

    for date_format in formats:
        try:
            result = datetime.strptime(
                text,
                date_format,
            )

            return datetime(
                result.year,
                result.month,
                result.day,
            )
        except ValueError:
            continue

    try:
        result = datetime.fromisoformat(
            text
        )

        return datetime(
            result.year,
            result.month,
            result.day,
        )
    except ValueError as error:
        raise ValueError(
            "format tanggal tidak valid: "
            f"{value}"
        ) from error


def display_date(
    value: datetime,
) -> str:
    return (
        f"{value.day:02d}/"
        f"{value.month:02d}/"
        f"{value.year}"
    )


def date_key(
    value: datetime,
) -> str:
    return (
        f"{value.year}"
        f"{value.month:02d}"
        f"{value.day:02d}"
    )


def with_time(
    value: datetime,
    hour: int,
    minute: int = 0,
) -> datetime:
    return datetime(
        value.year,
        value.month,
        value.day,
        hour,
        minute,
        0,
    )


def get_sheet(workbook):
    if PREFERRED_SHEET in workbook.sheetnames:
        return workbook[
            PREFERRED_SHEET
        ]

    return workbook[
        workbook.sheetnames[0]
    ]


def build_header_index(
    sheet,
) -> dict[str, int]:
    raw_headers = [
        cell.value
        for cell in sheet[1]
    ]

    headers = [
        normalize_header(value)
        for value in raw_headers
    ]

    aliases = {
        "date": [
            "tanggal",
            "date",
            "tgl",
        ],
        "product": [
            "merkjenisberas",
            "merkberas",
            "jenisberas",
            "produk",
            "namaproduk",
            "barang",
        ],
        "stock_in": [
            "stokmasuk",
            "stockin",
            "masuk",
            "qtymasuk",
            "jumlahmasuk",
        ],
        "stock_out": [
            "stokkeluar",
            "stockout",
            "keluar",
            "qtykeluar",
            "jumlahkeluar",
        ],
        "unit": [
            "satuan",
            "unit",
        ],
        "notes": [
            "catatan",
            "notes",
            "keterangan",
        ],
    }

    result: dict[str, int] = {}

    for key, possible_names in aliases.items():
        for possible_name in possible_names:
            if possible_name in headers:
                result[key] = headers.index(
                    possible_name
                )

                break

    required = [
        "date",
        "product",
        "stock_in",
        "stock_out",
    ]

    missing = [
        key
        for key in required
        if key not in result
    ]

    if missing:
        readable_headers = [
            str(
                value or ""
            ).strip()
            for value in raw_headers
        ]

        raise ValueError(
            "Kolom wajib tidak ditemukan: "
            f"{missing}. "
            "Header terbaca: "
            f"{readable_headers}"
        )

    return result


def get_product_code(
    product_id: str,
    fallback_codes: dict[str, str],
) -> str:
    if product_id in PRODUCT_CODES:
        return PRODUCT_CODES[
            product_id
        ]

    if product_id not in fallback_codes:
        next_number = (
            11
            + len(fallback_codes)
        )

        fallback_codes[
            product_id
        ] = (
            f"BR-{next_number:03d}"
        )

    return fallback_codes[
        product_id
    ]


def read_workbook(
    filename: str,
    expected_year: int,
    expected_month: int,
    source_order: int,
    fallback_codes: dict[str, str],
) -> list[RowData]:
    path = BASE_DIR / filename

    if not path.exists():
        raise FileNotFoundError(
            "File tidak ditemukan:\n"
            f"{path}"
        )

    workbook = openpyxl.load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    try:
        sheet = get_sheet(
            workbook
        )

        index = build_header_index(
            sheet
        )

        source_slug = slugify(
            path.stem
        )

        rows: list[RowData] = []

        for row_number, values in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            raw_date = values[
                index["date"]
            ]

            raw_product = values[
                index["product"]
            ]

            raw_stock_in = values[
                index["stock_in"]
            ]

            raw_stock_out = values[
                index["stock_out"]
            ]

            if (
                raw_date is None
                and raw_product is None
                and raw_stock_in is None
                and raw_stock_out is None
            ):
                continue

            try:
                item_date = parse_date(
                    raw_date
                )

                stock_in = parse_number(
                    raw_stock_in
                )

                stock_out = parse_number(
                    raw_stock_out
                )
            except Exception as error:
                raise ValueError(
                    f"{filename} "
                    f"baris {row_number}: "
                    f"{error}"
                ) from error

            if (
                item_date.year
                != expected_year
                or item_date.month
                != expected_month
            ):
                raise ValueError(
                    f"{filename} "
                    f"baris {row_number}: "
                    "tanggal "
                    f"{display_date(item_date)} "
                    "tidak sesuai bulan file "
                    f"{expected_month:02d}/"
                    f"{expected_year}"
                )

            if (
                stock_in < 0
                or stock_out < 0
            ):
                raise ValueError(
                    f"{filename} "
                    f"baris {row_number}: "
                    "stok tidak boleh negatif"
                )

            if (
                stock_in == 0
                and stock_out == 0
            ):
                continue

            product_id = slugify(
                raw_product
            )

            if not product_id:
                raise ValueError(
                    f"{filename} "
                    f"baris {row_number}: "
                    "nama produk kosong"
                )

            if "unit" in index:
                unit = str(
                    values[
                        index["unit"]
                    ]
                    or "Karung"
                ).strip()
            else:
                unit = "Karung"

            if not unit:
                unit = "Karung"

            if "notes" in index:
                notes = str(
                    values[
                        index["notes"]
                    ]
                    or ""
                ).strip()
            else:
                notes = ""

            rows.append(
                RowData(
                    source_file=filename,
                    source_order=source_order,
                    row_number=row_number,
                    row_key=(
                        f"{source_slug}_"
                        f"{row_number}"
                    ),
                    date=item_date,
                    product_id=product_id,
                    product_name=title_case(
                        raw_product
                    ),
                    product_code=(
                        get_product_code(
                            product_id,
                            fallback_codes,
                        )
                    ),
                    stock_in=stock_in,
                    stock_out=stock_out,
                    unit=unit,
                    notes=notes,
                )
            )

        return rows
    finally:
        workbook.close()


def read_all_rows() -> list[RowData]:
    rows: list[RowData] = []

    fallback_codes: dict[
        str,
        str,
    ] = {}

    for source_order, source in enumerate(
        EXCEL_FILES,
        start=1,
    ):
        filename, year, month = source

        print(
            f"Membaca {filename}..."
        )

        items = read_workbook(
            filename=filename,
            expected_year=year,
            expected_month=month,
            source_order=source_order,
            fallback_codes=fallback_codes,
        )

        print(
            f"  {len(items)} baris valid"
        )

        rows.extend(
            items
        )

    rows.sort(
        key=lambda item: (
            item.date,
            item.source_order,
            item.row_number,
        )
    )

    if not rows:
        raise ValueError(
            "Tidak ada data valid pada "
            "seluruh file Excel"
        )

    return rows


def add_keyword(
    result: set[str],
    value,
) -> None:
    text = str(
        value or ""
    ).strip().lower()

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    if (
        not text
        or len(text) > 250
    ):
        return

    result.add(
        text
    )

    compact = text.replace(
        " ",
        "",
    )

    if len(compact) >= 2:
        result.add(
            compact
        )

    words = text.split()[:30]

    for word in words:
        result.add(
            word
        )

        maximum_length = min(
            len(word),
            24,
        )

        for length in range(
            2,
            maximum_length + 1,
        ):
            result.add(
                word[:length]
            )


def build_search_keywords(
    batch: dict,
) -> list[str]:
    result: set[str] = set()

    main_values = (
        batch["id"],
        batch["batchCode"],
        batch["qrCodeValue"],
        batch["productId"],
        batch["productName"],
        batch["productCode"],
        batch["storageLocation"],
        batch["createdBy"],
        batch["createdByName"],
        batch["unit"],
        batch["initialQty"],
        batch["remainingQty"],
        batch["notes"],
    )

    for value in main_values:
        add_keyword(
            result,
            value,
        )

    if (
        batch["status"] == "active"
        and batch["remainingQty"] > 0
    ):
        status_values = (
            "active",
            "aktif",
            "batch aktif",
        )
    else:
        status_values = (
            "empty",
            "habis",
            "tidak aktif",
            "batch habis",
        )

    for value in status_values:
        add_keyword(
            result,
            value,
        )

    month_names = [
        "januari",
        "februari",
        "maret",
        "april",
        "mei",
        "juni",
        "juli",
        "agustus",
        "september",
        "oktober",
        "november",
        "desember",
    ]

    received_at = batch[
        "receivedAt"
    ]

    month_name = month_names[
        received_at.month - 1
    ]

    date_values = (
        display_date(
            received_at
        ),
        (
            f"{received_at.year}-"
            f"{received_at.month:02d}-"
            f"{received_at.day:02d}"
        ),
        (
            f"{received_at.day} "
            f"{month_name} "
            f"{received_at.year}"
        ),
        month_name,
        received_at.year,
    )

    for value in date_values:
        add_keyword(
            result,
            value,
        )

    return sorted(
        keyword
        for keyword in result
        if keyword
        and len(keyword) <= 250
    )


def allocate_location(
    available: list[str],
    occupied: dict[str, str],
    batch_code: str,
) -> str:
    if not available:
        active_locations = ", ".join(
            sorted(
                occupied,
                key=LOCATIONS.index,
            )
        )

        raise ValueError(
            "Kapasitas lokasi penyimpanan "
            "tidak cukup.\n"
            "Batch yang gagal dibuat: "
            f"{batch_code}\n"
            "Lokasi aktif: "
            f"{active_locations}"
        )

    return available.pop(0)


def release_location(
    location: str,
    available: list[str],
    occupied: dict[str, str],
) -> None:
    occupied.pop(
        location,
        None,
    )

    if location not in available:
        available.append(
            location
        )

        available.sort(
            key=LOCATIONS.index
        )


def build_data(
    rows: list[RowData],
) -> dict:
    products: dict[str, dict] = {}
    batches: dict[str, dict] = {}
    transactions: dict[str, dict] = {}

    queues: dict[
        str,
        list[QueueBatch],
    ] = {}

    sequences: dict[str, int] = {}

    available_locations = list(
        LOCATIONS
    )

    occupied_locations: dict[
        str,
        str,
    ] = {}

    stock_in_count = 0
    stock_out_count = 0

    for row in rows:
        if row.product_id not in products:
            products[
                row.product_id
            ] = {
                "id": row.product_id,
                "name": row.product_name,
                "code": row.product_code,
                "category": "Beras",
                "unit": row.unit,
                "minimumStock": (
                    DEFAULT_MINIMUM_STOCK
                ),
                "minStock": (
                    DEFAULT_MINIMUM_STOCK
                ),
                "totalStock": 0,
                "isActive": True,
                "isDeleted": False,
                "createdAt": (
                    firestore.SERVER_TIMESTAMP
                ),
                "updatedAt": (
                    firestore.SERVER_TIMESTAMP
                ),
            }

            queues[
                row.product_id
            ] = []

            sequences[
                row.product_id
            ] = 0

        if row.stock_in > 0:
            sequences[
                row.product_id
            ] += 1

            sequence = sequences[
                row.product_id
            ]

            batch_code = (
                f"BATCH-"
                f"{date_key(row.date)}-"
                f"{row.product_code}-"
                f"{sequence:03d}"
            )

            location = allocate_location(
                available=available_locations,
                occupied=occupied_locations,
                batch_code=batch_code,
            )

            occupied_locations[
                location
            ] = batch_code

            batches[
                batch_code
            ] = {
                "id": batch_code,
                "productId": (
                    row.product_id
                ),
                "productName": (
                    row.product_name
                ),
                "productCode": (
                    row.product_code
                ),
                "batchCode": batch_code,
                "receivedAt": with_time(
                    row.date,
                    8,
                ),
                "initialQty": (
                    row.stock_in
                ),
                "remainingQty": (
                    row.stock_in
                ),
                "unit": row.unit,
                "qrCodeValue": (
                    batch_code
                ),
                "status": "active",
                "storageLocation": (
                    location
                ),
                "createdBy": (
                    IMPORT_USER_ID
                ),
                "createdByName": (
                    IMPORT_USER_NAME
                ),
                "notes": (
                    "Import stok masuk dari "
                    f"{row.source_file}, "
                    "tanggal "
                    f"{display_date(row.date)}. "
                    f"{row.notes}"
                ).strip(),
                "createdAt": with_time(
                    row.date,
                    8,
                ),
                "updatedAt": (
                    firestore.SERVER_TIMESTAMP
                ),
            }

            queues[
                row.product_id
            ].append(
                QueueBatch(
                    batch_id=batch_code,
                    remaining_qty=(
                        row.stock_in
                    ),
                    location=location,
                )
            )

            transaction_id = (
                f"tx_in_"
                f"{row.row_key}_"
                f"{batch_code}"
            )

            transactions[
                transaction_id
            ] = {
                "id": transaction_id,
                "type": "stock_in",
                "productId": (
                    row.product_id
                ),
                "productName": (
                    row.product_name
                ),
                "batchId": batch_code,
                "batchCode": batch_code,
                "qty": row.stock_in,
                "unit": row.unit,
                "performedBy": (
                    IMPORT_USER_ID
                ),
                "performedByName": (
                    IMPORT_USER_NAME
                ),
                "notes": (
                    "Import stok masuk dari "
                    f"{row.source_file}, "
                    "tanggal "
                    f"{display_date(row.date)}. "
                    f"{row.notes}"
                ).strip(),
                "createdAt": with_time(
                    row.date,
                    8,
                    10,
                ),
            }

            stock_in_count += 1

        remaining_out = row.stock_out
        part_number = 1

        while remaining_out > 0:
            current_batch = next(
                (
                    item
                    for item
                    in queues[
                        row.product_id
                    ]
                    if item.remaining_qty > 0
                ),
                None,
            )

            if current_batch is None:
                raise ValueError(
                    "Stok keluar melebihi "
                    "stok tersedia.\n"
                    f"File: {row.source_file}\n"
                    f"Baris: {row.row_number}\n"
                    "Tanggal: "
                    f"{display_date(row.date)}\n"
                    f"Produk: {row.product_name}\n"
                    "Kekurangan: "
                    f"{remaining_out} "
                    f"{row.unit}"
                )

            allocated_quantity = min(
                remaining_out,
                current_batch
                .remaining_qty,
            )

            current_batch.remaining_qty -= (
                allocated_quantity
            )

            remaining_out -= (
                allocated_quantity
            )

            batch = batches[
                current_batch.batch_id
            ]

            batch[
                "remainingQty"
            ] = current_batch.remaining_qty

            if (
                current_batch
                .remaining_qty
                <= 0
            ):
                batch["status"] = "empty"
            else:
                batch["status"] = "active"

            batch[
                "updatedAt"
            ] = firestore.SERVER_TIMESTAMP

            if (
                current_batch
                .remaining_qty
                <= 0
            ):
                release_location(
                    location=(
                        current_batch.location
                    ),
                    available=(
                        available_locations
                    ),
                    occupied=(
                        occupied_locations
                    ),
                )

            transaction_id = (
                f"tx_out_"
                f"{row.row_key}_"
                f"{part_number:03d}_"
                f"{current_batch.batch_id}"
            )

            transactions[
                transaction_id
            ] = {
                "id": transaction_id,
                "type": "stock_out",
                "productId": (
                    row.product_id
                ),
                "productName": (
                    row.product_name
                ),
                "batchId": (
                    current_batch.batch_id
                ),
                "batchCode": (
                    current_batch.batch_id
                ),
                "qty": (
                    allocated_quantity
                ),
                "unit": row.unit,
                "performedBy": (
                    IMPORT_USER_ID
                ),
                "performedByName": (
                    IMPORT_USER_NAME
                ),
                "notes": (
                    "Import stok keluar FIFO "
                    f"dari {row.source_file}, "
                    "tanggal "
                    f"{display_date(row.date)}. "
                    f"{row.notes}"
                ).strip(),
                "createdAt": with_time(
                    row.date,
                    15,
                    min(
                        part_number,
                        59,
                    ),
                ),
            }

            stock_out_count += 1
            part_number += 1

    for product_id, product in products.items():
        product[
            "totalStock"
        ] = sum(
            max(
                0,
                int(
                    batch[
                        "remainingQty"
                    ]
                ),
            )
            for batch in batches.values()
            if (
                batch["productId"]
                == product_id
            )
        )

    for batch in batches.values():
        batch[
            "searchKeywords"
        ] = build_search_keywords(
            batch
        )

    counters = {
        (
            f"batch_sequence_"
            f"{product_id}"
        ): {
            "id": (
                f"batch_sequence_"
                f"{product_id}"
            ),
            "productId": product_id,
            "productCode": (
                products[
                    product_id
                ]["code"]
            ),
            "lastNumber": last_number,
            "updatedAt": (
                firestore.SERVER_TIMESTAMP
            ),
        }
        for product_id, last_number
        in sequences.items()
    }

    active_by_location = {
        batch["storageLocation"]: batch
        for batch in batches.values()
        if (
            batch["status"] == "active"
            and int(
                batch["remainingQty"]
            ) > 0
        )
    }

    locations: dict[
        str,
        dict,
    ] = {}

    for location_code in LOCATIONS:
        active_batch = (
            active_by_location.get(
                location_code
            )
        )

        locations[
            location_code
        ] = {
            "id": location_code,
            "locationCode": (
                location_code
            ),
            "isOccupied": (
                active_batch is not None
            ),
            "batchId": (
                active_batch["id"]
                if active_batch
                else ""
            ),
            "batchCode": (
                active_batch["batchCode"]
                if active_batch
                else ""
            ),
            "productId": (
                active_batch["productId"]
                if active_batch
                else ""
            ),
            "productName": (
                active_batch["productName"]
                if active_batch
                else ""
            ),
            "remainingQty": (
                int(
                    active_batch[
                        "remainingQty"
                    ]
                )
                if active_batch
                else 0
            ),
            "occupiedAt": (
                active_batch["receivedAt"]
                if active_batch
                else None
            ),
            "updatedAt": (
                firestore.SERVER_TIMESTAMP
            ),
        }

    total_stock_in = sum(
        row.stock_in
        for row in rows
    )

    total_stock_out = sum(
        row.stock_out
        for row in rows
    )

    final_stock = sum(
        product["totalStock"]
        for product in products.values()
    )

    expected_final_stock = (
        total_stock_in
        - total_stock_out
    )

    if final_stock != expected_final_stock:
        raise RuntimeError(
            "Validasi stok gagal.\n"
            f"Total masuk : {total_stock_in}\n"
            f"Total keluar: {total_stock_out}\n"
            "Seharusnya  : "
            f"{expected_final_stock}\n"
            f"Hasil FIFO  : {final_stock}"
        )

    return {
        "rows": len(rows),
        "products": products,
        "batches": batches,
        "transactions": transactions,
        "counters": counters,
        "locations": locations,
        "stock_in_count": (
            stock_in_count
        ),
        "stock_out_count": (
            stock_out_count
        ),
        "total_in": (
            total_stock_in
        ),
        "total_out": (
            total_stock_out
        ),
        "final_stock": (
            final_stock
        ),
    }


def print_summary(
    data: dict,
) -> None:
    print()
    print("RINGKASAN IMPORT")
    print("=================")
    print(
        "Periode                 : "
        "Januari-Juli 2026"
    )
    print(
        "Baris diproses          : "
        f"{data['rows']}"
    )
    print(
        "Produk                  : "
        f"{len(data['products'])}"
    )
    print(
        "Batch                   : "
        f"{len(data['batches'])}"
    )
    print(
        "Transaksi stock_in      : "
        f"{data['stock_in_count']}"
    )
    print(
        "Transaksi stock_out     : "
        f"{data['stock_out_count']}"
    )
    print(
        "Counter                 : "
        f"{len(data['counters'])}"
    )
    print(
        "Lokasi penyimpanan      : "
        f"{len(data['locations'])}"
    )
    print(
        "Total stok masuk        : "
        f"{data['total_in']}"
    )
    print(
        "Total stok keluar       : "
        f"{data['total_out']}"
    )
    print(
        "Sisa stok akhir         : "
        f"{data['final_stock']}"
    )

    print()
    print("PRODUK")
    print("======")

    products = sorted(
        data["products"].values(),
        key=lambda item: item["code"],
    )

    for product in products:
        print(
            f"{product['code']} | "
            f"{product['name']} | "
            "stok akhir: "
            f"{product['totalStock']} "
            f"{product['unit']}"
        )


def append_documents(
    writes: list,
    database,
    collection_name: str,
    documents: dict,
) -> None:
    for document_id, value in (
        documents.items()
    ):
        reference = (
            database
            .collection(collection_name)
            .document(document_id)
        )

        writes.append(
            (
                reference,
                value,
            )
        )


def commit_data(
    database,
    data: dict,
) -> None:
    writes: list = []

    append_documents(
        writes,
        database,
        COL_PRODUCTS,
        data["products"],
    )

    append_documents(
        writes,
        database,
        COL_BATCHES,
        data["batches"],
    )

    append_documents(
        writes,
        database,
        COL_TRANSACTIONS,
        data["transactions"],
    )

    append_documents(
        writes,
        database,
        COL_COUNTERS,
        data["counters"],
    )

    append_documents(
        writes,
        database,
        COL_LOCATIONS,
        data["locations"],
    )

    print()
    print(
        "Total write ke Firestore: "
        f"{len(writes)}"
    )

    chunk_size = 400

    for start in range(
        0,
        len(writes),
        chunk_size,
    ):
        chunk = writes[
            start:
            start + chunk_size
        ]

        write_batch = database.batch()

        for reference, value in chunk:
            write_batch.set(
                reference,
                value,
            )

        write_batch.commit()

        print(
            f"Commit {start + 1}-"
            f"{start + len(chunk)} "
            "selesai"
        )


def write_import_log(
    database,
    data: dict,
) -> None:
    source_files = [
        filename
        for filename, _, _
        in EXCEL_FILES
    ]

    database.collection(
        COL_IMPORT_LOGS
    ).document(
        IMPORT_LOG_ID
    ).set(
        {
            "id": IMPORT_LOG_ID,
            "description": (
                "Import data awal stok beras "
                "Januari-Juli 2026"
            ),
            "periodStart": (
                datetime(2026, 1, 1)
            ),
            "periodEnd": (
                datetime(2026, 7, 31)
            ),
            "sourceFiles": source_files,
            "rowsProcessed": (
                data["rows"]
            ),
            "productsCount": len(
                data["products"]
            ),
            "batchesCount": len(
                data["batches"]
            ),
            "stockInTransactionCount": (
                data["stock_in_count"]
            ),
            "stockOutTransactionCount": (
                data["stock_out_count"]
            ),
            "countersCount": len(
                data["counters"]
            ),
            "storageLocationsCount": len(
                data["locations"]
            ),
            "totalStockIn": (
                data["total_in"]
            ),
            "totalStockOut": (
                data["total_out"]
            ),
            "finalStock": (
                data["final_stock"]
            ),
            "createdAt": (
                firestore.SERVER_TIMESTAMP
            ),
        }
    )


def count_documents(
    database,
    collection_name: str,
) -> int:
    return sum(
        1
        for _ in database
        .collection(collection_name)
        .stream()
    )


def verify_import(
    database,
    data: dict,
) -> None:
    expected_counts = {
        COL_PRODUCTS: len(
            data["products"]
        ),
        COL_BATCHES: len(
            data["batches"]
        ),
        COL_TRANSACTIONS: len(
            data["transactions"]
        ),
        COL_COUNTERS: len(
            data["counters"]
        ),
        COL_LOCATIONS: len(
            data["locations"]
        ),
        COL_IMPORT_LOGS: 1,
    }

    errors: list[str] = []

    print()
    print("VERIFIKASI FIRESTORE")
    print("====================")

    for (
        collection_name,
        expected_count,
    ) in expected_counts.items():
        actual_count = count_documents(
            database,
            collection_name,
        )

        print(
            f"- {collection_name}: "
            f"{actual_count} dokumen"
        )

        if actual_count != expected_count:
            errors.append(
                f"{collection_name}: "
                f"aktual {actual_count}, "
                "seharusnya "
                f"{expected_count}"
            )

    for product_id, product in (
        data["products"].items()
    ):
        snapshot = (
            database
            .collection(COL_PRODUCTS)
            .document(product_id)
            .get()
        )

        if snapshot.exists:
            saved_data = (
                snapshot.to_dict()
                or {}
            )
        else:
            saved_data = {}

        actual_stock = int(
            saved_data.get(
                "totalStock",
                -1,
            )
        )

        expected_stock = int(
            product["totalStock"]
        )

        if actual_stock != expected_stock:
            errors.append(
                f"stok {product_id}: "
                f"aktual {actual_stock}, "
                "seharusnya "
                f"{expected_stock}"
            )

    if errors:
        raise RuntimeError(
            "Verifikasi gagal:\n- "
            + "\n- ".join(errors)
        )

    print()
    print(
        "IMPORT BERHASIL DAN TERVERIFIKASI"
    )
    print(
        "================================="
    )
    print(
        "Data Januari-Juli 2026 "
        "sudah masuk."
    )
    print(
        "Collection users tidak disentuh."
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import stok Januari-Juli 2026 "
            "ke Cloud Firestore"
        )
    )

    parser.add_argument(
        "--commit",
        action="store_true",
        help=(
            "Benar-benar menulis data "
            "ke Firestore"
        ),
    )

    parser.add_argument(
        "--force",
        action="store_true",
        help=(
            "Abaikan import log lama"
        ),
    )

    arguments = parser.parse_args()

    print()
    print(
        "MODE:",
        (
            "COMMIT KE FIRESTORE"
            if arguments.commit
            else "DRY RUN"
        ),
    )
    print()

    rows = read_all_rows()

    data = build_data(
        rows
    )

    print_summary(
        data
    )

    if not arguments.commit:
        print()
        print("DRY RUN BERHASIL")
        print("================")
        print(
            "Belum ada data yang ditulis."
        )
        print(
            "Setelah cleanup, jalankan:"
        )
        print(
            "python import_initial_stock_2026.py "
            "--commit"
        )

        return

    database = initialize_firestore()

    import_log = (
        database
        .collection(COL_IMPORT_LOGS)
        .document(IMPORT_LOG_ID)
        .get()
    )

    if (
        import_log.exists
        and not arguments.force
    ):
        raise RuntimeError(
            "Import Januari-Juli "
            "sudah pernah dilakukan. "
            "Jalankan cleanup dahulu "
            "atau gunakan --force."
        )

    print()
    print("MULAI IMPORT")
    print("============")

    commit_data(
        database,
        data,
    )

    write_import_log(
        database,
        data,
    )

    verify_import(
        database,
        data,
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print()
        print(
            "Proses dihentikan."
        )

        raise SystemExit(130)
    except Exception as error:
        print()
        print("IMPORT GAGAL")
        print("============")
        print(error)

        sys.exit(1)